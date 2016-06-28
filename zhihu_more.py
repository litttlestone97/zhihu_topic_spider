#-*- coding:utf-8 -*-
#加载number后面18条话题
import urllib
import urllib.parse
import urllib.error
import urllib.request
import json
import re
from openpyxl import load_workbook
from openpyxl import Workbook 

def zhihu_topic_more(number):

    #打开文件
    wb = load_workbook("know_topic.xlsx")
    ws=wb["topic"]

    #获取连接
    url = 'https://m.zhihu.com/node/TopicsPlazzaListV2'
    user_agent = 'Mozilla/4.0 (compatible; MSIE 5.5; Windows NT)'
    params=json.dumps({"topic_id":253,"offset":number,"hash_id":"b141d89903cd6c510bca526366060d47"})
    data = {'method':'next','params':params,'_xsrf':'c50ef9467598f0c770cb96df57891f43'}
    data = urllib.parse.urlencode(data)
    data = data.encode('utf-8') #编译   
    try:
        request = urllib.request.Request(url=url,data=data) 
        response = urllib.request.urlopen(request)
        #解析json
        result = json.loads(response.read().decode('utf-8'))
        
        print ('获取第',number,'条数据成功')
        content = result['msg']
        #匹配链接
        #循环取strong/href/p/ href="/topic/
        topic='<strong>(.*)</strong>'
        meaning='<p>(.*)</p>'
        href='href="(/topic/.*)">'

        href_list = re.compile('href="(/topic/.*)">')
        meaning_list = re.compile('<p>(.*)</p>')
        topic_list = re.compile('<strong>(.*)</strong>')

        for x in range(0, 19):
            topic_items = re.findall(topic_list,content[x])
            meaning_items = re.findall(meaning_list,content[x])
            href_items = re.findall(href_list,content[x])
            #打印
            #print (i,topic_items,meaning_items,href_items)
            ws.cell(row = number+x+2,column= 1).value=number+x
            ws.cell(row = number+x+2,column= 2).value=topic_items[0]
            ws.cell(row = number+x+2,column= 3).value=meaning_items[0]
            ws.cell(row = number+x+2,column= 4).value='https://www.zhihu.com'+href_items[0]
            
            wb.save(filename = "know_topic.xlsx")   

    except urllib.error.URLError as e:
        if hasattr(e,"code"):
            print (e.code)
        if hasattr(e,"reason"):
            print (e.reason)

    return