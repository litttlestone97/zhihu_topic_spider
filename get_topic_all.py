#-*- coding:utf-8 -*-
#刷新话题信息

'''
功能描述：
+ 话题关注人数  know_topic.xlsx
+ 更新话题描述  know_topic.xlsx

编写日期：20160702
编写人：little_Stone
'''

import urllib
import urllib.parse
import urllib.error
import urllib.request
import json
import re
from openpyxl import load_workbook
from openpyxl import Workbook 
from function import *    #模块识别失效时。采用此方式


wb = load_workbook("know_topic.xlsx")
sheet_names = wb.get_sheet_names() 
#打开文件
for index in range(32,33):

	ws = wb.get_sheet_by_name(sheet_names[index])# index为0为第一张表
	rows=ws.max_row   #获取行数

	ws.cell(row = 1,column= 5).value="关注人数"

	#更新话题描述，话题关注人数   rows-2
	for i in range(0,rows-2):	
		href= ws.cell(row = i+2,column= 4).value
		#print (href,'i=',i)
		(meaning,people_number)=topic_all(href)
		#print(i+2,meaning,'meaning长度为',len(meaning))
		if len(meaning)!=0:
			ws.cell(row =i+2,column= 3).value=meaning[0]
		if len(people_number)!=0:
			ws.cell(row =i+2,column= 5).value=people_number[0]
		if i%200 ==0 :
			print('正在写入第',i,'条数据')
	print('完成第',index+1,'页数据','本页',ws.max_row,'条数据\n')
wb.save(filename = "know_topic.xlsx")