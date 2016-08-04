#-*- coding:utf-8 -*- 
'''
功能描述：
+ 按照主话题列表，查询字话题，并写入

编写日期：20160702
编写人：little_Stone
'''
 

import function
from openpyxl import load_workbook
from openpyxl import Workbook 

#打开文件
wb2 = load_workbook("main_topic.xlsx")
ws2 = wb2["主话题列表"] 
ws2.cell(row = 1,column= 4).value="子话题数量"

##必须先建立文件

for a in range(2,35):#共33个主话题,从第a-1个话题开始

 	#新建sheet	
	wb1 = load_workbook("know_topic.xlsx")
	ws1 = wb1.create_sheet()	
	#设置ws的名称
	ws1.title = ws2.cell(row = a,column= 2).value
	ws1.cell(row = 1,column= 1).value='编号'
	ws1.cell(row = 1,column= 2).value='话题'
	ws1.cell(row = 1,column= 3).value='简介'
	ws1.cell(row = 1,column= 4).value='链接' 
	wb1.save(filename = "know_topic.xlsx")


	#加载话题18*j,覆盖加载
	j =0
	#断点跑数据时，修改j的初始值   id-1

	sheet = ws2.cell(row = a,column= 2).value  #主话题
	topic_id=ws2.cell(row = a,column= 3).value  #主话题id

	print(a-1,sheet,'ID=',topic_id,'j=',j)

	#循环请求话题数据，到达19万时，手动重复
	for i in range(0,10000):	 
		i+=1
		step=function.topic_more(j,sheet,topic_id)
		j+=step
		if step == 0:
			break
	print (sheet,'共',j,'条子话题，over\n')
	ws2.cell(row = a,column= 4).value=j
wb2.save(filename = "main_topic.xlsx")