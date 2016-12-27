#-*-coding:utf8-*-
'''
功能描述：
+ 抓取知乎主话题,单页

编写日期：20160702
编写人：little_Stone
'''

import requests
import math
from openpyxl import load_workbook
from openpyxl import Workbook 
import re      #正则表达式的库

#hea是我们自己构造的一个字典，里面保存了user-agent
hea = {'User-Agent':'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2272.118 Safari/537.36'}
html = requests.get('https://www.zhihu.com/topics',headers = hea)
#print (html.text)
#写入文档
fp = open('know_topic.txt','w',encoding='utf-8')
fp.write(html.text)
fp.close()


#爬取话题标题及连接
 	#创建一个workbook
    #sheet=wb.add_sheet("xlwt3数据测试表")
wb=Workbook()
dest_filename = 'main_topic.xlsx'
#第一个sheet是ws
ws = wb.worksheets[0]
#ws=wb.create_sheet()
#设置ws的名称
ws.title = u"主话题列表"
ws.cell(row = 1,column= 1).value='编号'
ws.cell(row = 1,column= 2).value='话题'
ws.cell(row = 1,column= 3).value='话题ID'

topic='<a href="#(.*)">'
topic_id='data-id="(.*)"><a href="#'


topic_list = re.findall(topic,html.text)
topic_id_list = re.findall(topic_id,html.text) 


j=2
for each in topic_list:
	ws.cell(row = j,column= 1).value=j-1
	ws.cell(row = j,column= 2).value=each
	j+=1
i=2
for each in topic_id_list:
	ws.cell(row = i,column= 3).value=each
	i += 1

wb.save(filename = dest_filename)	    
print("main_topic over")
