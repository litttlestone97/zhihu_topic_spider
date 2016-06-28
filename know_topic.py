#抓取知乎话题,单页

#-*-coding:utf8-*-
import requests
import math
from openpyxl import load_workbook
from openpyxl import Workbook 
import re      #正则表达式的库

#hea是我们自己构造的一个字典，里面保存了user-agent
hea = {'User-Agent':'Mozilla/5.0 (Windows NT 6.3; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2272.118 Safari/537.36'}
html = requests.get('https://www.zhihu.com/topics',headers = hea)
#print (html.text)
#写入文档
fp = open('know_topic.txt','w',encoding='utf-8')
fp.write(html.text)
fp.close()


#爬取话题标题及连接

 	#创建一个workbook
    #sheet=wb.add_sheet("xlwt3数据测试表")
wb=Workbook()
dest_filename = 'know_topic.xlsx'
#第一个sheet是ws
ws = wb.worksheets[0]
#ws=wb.create_sheet()
#设置ws的名称
ws.title = u"话题列表"
ws.cell(row = 1,column= 1).value='编号'
ws.cell(row = 1,column= 2).value='话题'
ws.cell(row = 1,column= 3).value='简介'
ws.cell(row = 1,column= 4).value='链接' 

#循环取strong/href/p/ href="/topic/
topic='<strong>(.*)</strong>'
meaning='<p>(.*)</p>'
href='href="(/topic/.*)">'


topic_list = re.findall(topic,html.text)
meaning_list = re.findall(meaning,html.text)
href_list = re.findall(href,html.text)


j=2
for each in topic_list:
	ws.cell(row = j,column= 1).value=j-1
	ws.cell(row = j,column= 2).value=each
	j+=1
i=2
for each in meaning_list:
	ws.cell(row = i,column= 3).value=each
	i += 1
i=2
for each in href_list:
	ws.cell(row = i,column= 4).value='https://www.zhihu.com'+each
	i += 1
	if i==j:
		print(i,j)
		break

wb.save(filename = dest_filename)	    
print("OK")
