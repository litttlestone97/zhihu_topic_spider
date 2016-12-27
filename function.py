#-*- coding:utf-8 -*-
'''
功能描述：函数库
编写日期：20160702
编写人：little_Stone
'''
import urllib
import urllib.parse
import urllib.error
import requests
import json
import re
#from bs4 import BeautifulSoup
import urllib.request
from openpyxl import load_workbook
from openpyxl import Workbook 

import requests,time
from bs4 import BeautifulSoup
import urllib
import urllib.parse
import urllib.error
import urllib.request
from openpyxl import load_workbook
from openpyxl import Workbook 

#验证码获取
def get_captcha(data):
    with open('captcha.gif','wb') as fp:
        fp.write(data)
    return input('输入验证码：')
 
#登陆
def login(username,password,oncaptcha):
    sessiona = requests.Session()
    headers = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:47.0) Gecko/20100101 Firefox/47.0'}
    _xsrf = BeautifulSoup(sessiona.get('https://www.zhihu.com/#signin',headers=headers).content,'html.parser').find('input',attrs={'name':'_xsrf'}).get('value')
    captcha_content = sessiona.get('https://www.zhihu.com/captcha.gif?r=%d&type=login'%(time.time()*1000),headers=headers).content
    data = {
        "_xsrf":_xsrf,
        "email":username,
        "password":password,
        "remember_me":True,
        "captcha":oncaptcha(captcha_content)
    }
    resp = sessiona.post('https://www.zhihu.com/login/email',data,headers=headers).content
    print(resp)
    return sessiona 


#根据输入链接，更新话题描述，话题关注人数
def topic_all(href):
	hea = {'User-Agent':'Mozilla/4.0 (compatible; MSIE 5.5; Windows NT)'}
	html = requests.get(href,headers = hea)

	#fp = open('1.txt','w',encoding='utf-8')
	#fp.write(html.text)
	#fp.close()

	#循环取话题描述，话题关注人数
	meaning_s='<div class="zm-editable-content" data-editable-maxlength="130" data-disabled="1">(.*)</div>'
	people_number_s='<strong>(.*)</strong> 人关注了该话题'

	meaning = re.findall(meaning_s,html.text)
	people_number = re.findall(people_number_s,html.text)

	#print(meaning,people_number)
	return (meaning,people_number)



#加载话题后面一个请求的数据，并写入文件(起始点，主话题sheet名，主话题id)
def topic_more(number,sheet,topic_id):

    #打开文件
    wb = load_workbook("know_topic.xlsx")
    ws=wb[sheet]

    #获取连接
    url = 'https://m.zhihu.com/node/TopicsPlazzaListV2'
    user_agent = 'Mozilla/4.0 (compatible; MSIE 5.5; Windows NT)'
    params=json.dumps({"topic_id":topic_id,"offset":number,"hash_id":"b141d89903cd6c510bca526366060d47"})
    data = {'method':'next','params':params,'_xsrf':'c50ef9467598f0c770cb96df57891f43'}
    data = urllib.parse.urlencode(data)
    data = data.encode('utf-8') #编译   
    try:
        request = urllib.request.Request(url=url,data=data) 
        response = urllib.request.urlopen(request)
        #解析json
        result = json.loads(response.read().decode('utf-8'))
        
        if number%100 ==0  :
            print ('正在写入第',number,'条数据')            
        content = result['msg']

        step=len(content)
        #print('step',step) #验证步长
        #匹配链接
        #循环取strong/href/p/ href="/topic/
        topic='<strong>(.*)</strong>'
        meaning='<p>(.*)</p>'
        href='href="(/topic/.*)">'

        href_list = re.compile('href="(/topic/.*)">')
        meaning_list = re.compile('<p>(.*)</p>')
        topic_list = re.compile('<strong>(.*)</strong>')


        for x in range(0, step):
            topic_items = re.findall(topic_list,content[x])
            meaning_items = re.findall(meaning_list,content[x])
            href_items = re.findall(href_list,content[x])
            #打印
            #Sprint (x,topic_items,href_items)
            ws.cell(row = number+x+2,column= 1).value=number+x
            ws.cell(row = number+x+2,column= 2).value=topic_items[0]
            #ws.cell(row = number+x+2,column= 3).value=meaning_items[0]
            ws.cell(row = number+x+2,column= 4).value='https://www.zhihu.com'+href_items[0]
        wb.save(filename = "know_topic.xlsx")  

    except urllib.error.URLError as e:
        if hasattr(e,"code"):
            print (e.code)
        if hasattr(e,"reason"):
            print (e.reason)

    return (step)